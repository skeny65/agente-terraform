#!/usr/bin/env python3
"""
leer_excel.py — Convierte un .xlsx en JSON (o CSV) SIN usar el LLM.

Primera fila = cabeceras. Cada fila siguiente = un dict {cabecera: valor}.
Usa openpyxl si está instalado; si no, un lector mínimo con librería estándar.

Uso:
    python leer_excel.py datos.xlsx -o datos.json
    python leer_excel.py datos.xlsx --hoja "VMs" --csv > vms.csv
    python leer_excel.py datos.xlsx --hojas          # lista las hojas y sale

100% offline.
"""
from __future__ import annotations

import argparse
import csv
import io
import json
import re
import sys
import zipfile
import xml.etree.ElementTree as ET
from pathlib import Path


# --- Lector con openpyxl (preferido) ---------------------------------------

def _leer_openpyxl(ruta: Path, hoja: str | None) -> dict[str, list[dict]]:
    import openpyxl  # type: ignore
    wb = openpyxl.load_workbook(ruta, read_only=True, data_only=True)
    hojas = [hoja] if hoja else wb.sheetnames
    out: dict[str, list[dict]] = {}
    for nombre in hojas:
        ws = wb[nombre]
        filas = ws.iter_rows(values_only=True)
        try:
            cabeceras = [str(c).strip() if c is not None else f"col{i+1}"
                         for i, c in enumerate(next(filas))]
        except StopIteration:
            out[nombre] = []
            continue
        registros = []
        for fila in filas:
            if fila is None or all(v is None or str(v).strip() == "" for v in fila):
                continue
            registros.append({cab: ("" if v is None else v) for cab, v in zip(cabeceras, fila)})
        out[nombre] = registros
    wb.close()
    return out


# --- Lector de reserva con solo stdlib -----------------------------------

_NS = {"m": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}


def _col_a_indice(ref: str) -> int:
    letras = re.match(r"[A-Z]+", ref).group(0)
    n = 0
    for ch in letras:
        n = n * 26 + (ord(ch) - 64)
    return n - 1


def _leer_stdlib(ruta: Path, hoja: str | None) -> dict[str, list[dict]]:
    with zipfile.ZipFile(ruta) as z:
        shared: list[str] = []
        if "xl/sharedStrings.xml" in z.namelist():
            raiz = ET.fromstring(z.read("xl/sharedStrings.xml"))
            for si in raiz.findall("m:si", _NS):
                shared.append("".join(t.text or "" for t in si.iter("{%s}t" % _NS["m"])))

        wb = ET.fromstring(z.read("xl/workbook.xml"))
        rels = ET.fromstring(z.read("xl/_rels/workbook.xml.rels"))
        rid_a_destino = {r.get("Id"): r.get("Target") for r in rels}
        hojas_def = []
        for s in wb.findall("m:sheets/m:sheet", _NS):
            rid = s.get("{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id")
            destino = rid_a_destino.get(rid, "")
            ruta_hoja = "xl/" + destino.lstrip("/") if not destino.startswith("xl/") else destino
            hojas_def.append((s.get("name"), ruta_hoja))

        objetivo = [(n, p) for n, p in hojas_def if (hoja is None or n == hoja)]
        out: dict[str, list[dict]] = {}
        for nombre, ruta_hoja in objetivo:
            data = ET.fromstring(z.read(ruta_hoja))
            matriz: list[list[str]] = []
            for fila in data.findall("m:sheetData/m:row", _NS):
                valores: dict[int, str] = {}
                for c in fila.findall("m:c", _NS):
                    idx = _col_a_indice(c.get("r", "A1"))
                    v = c.find("m:v", _NS)
                    txt = ""
                    if v is not None and v.text is not None:
                        txt = shared[int(v.text)] if c.get("t") == "s" else v.text
                    elif c.find("m:is", _NS) is not None:
                        txt = "".join(t.text or "" for t in c.find("m:is", _NS).iter("{%s}t" % _NS["m"]))
                    valores[idx] = txt
                ancho = (max(valores) + 1) if valores else 0
                matriz.append([valores.get(i, "") for i in range(ancho)])
            if not matriz:
                out[nombre] = []
                continue
            ancho = max(len(f) for f in matriz)
            cabeceras = [(matriz[0][i].strip() if i < len(matriz[0]) and matriz[0][i] else f"col{i+1}") for i in range(ancho)]
            registros = []
            for fila in matriz[1:]:
                fila = fila + [""] * (ancho - len(fila))
                if all(str(x).strip() == "" for x in fila):
                    continue
                registros.append(dict(zip(cabeceras, fila)))
            out[nombre] = registros
        return out


def leer(ruta: Path, hoja: str | None = None) -> dict[str, list[dict]]:
    try:
        return _leer_openpyxl(ruta, hoja)
    except ImportError:
        return _leer_stdlib(ruta, hoja)


# --- CLI -----------------------------------------------------------------

def _utf8_stdio() -> None:
    for flujo in (sys.stdout, sys.stderr):
        try:
            flujo.reconfigure(encoding="utf-8", errors="replace")
        except (AttributeError, ValueError):
            pass


def main(argv=None) -> int:
    _utf8_stdio()
    ap = argparse.ArgumentParser(description="Convierte un .xlsx en JSON/CSV (sin LLM).")
    ap.add_argument("xlsx", type=Path)
    ap.add_argument("-o", "--salida", type=Path, help="JSON de salida (por defecto: stdout)")
    ap.add_argument("--hoja", help="Nombre de una hoja concreta (por defecto: todas)")
    ap.add_argument("--hojas", action="store_true", help="Solo listar los nombres de las hojas")
    ap.add_argument("--csv", action="store_true", help="Emitir CSV (requiere --hoja o que haya una sola)")
    a = ap.parse_args(argv)

    if not a.xlsx.is_file():
        print(f"error: no existe {a.xlsx}", file=sys.stderr)
        return 2

    try:
        datos = leer(a.xlsx, a.hoja)
    except (zipfile.BadZipFile, ET.ParseError, KeyError) as e:
        print(f"error al leer el xlsx: {e}", file=sys.stderr)
        return 1

    if a.hojas:
        for n, filas in datos.items():
            print(f"{n}\t{len(filas)} filas")
        return 0

    if a.csv:
        if len(datos) != 1:
            print("error: --csv necesita --hoja (hay varias hojas)", file=sys.stderr)
            return 2
        filas = next(iter(datos.values()))
        if not filas:
            return 0
        w = csv.DictWriter(sys.stdout, fieldnames=list(filas[0].keys()))
        w.writeheader()
        w.writerows(filas)
        return 0

    payload = {"fuente": a.xlsx.name, "hojas": {n: filas for n, filas in datos.items()}}
    texto = json.dumps(payload, ensure_ascii=False, indent=2, default=str)
    if a.salida:
        a.salida.write_text(texto, encoding="utf-8")
        total = sum(len(f) for f in datos.values())
        print(f"escrito: {a.salida}  ({len(datos)} hoja(s), {total} fila(s))", file=sys.stderr)
    else:
        print(texto)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
